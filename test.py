from selenium import webdriver
from bs4 import BeautifulSoup
from time import sleep
import openpyxl


driver = webdriver.Firefox(executable_path=r"webdrivers\\geckodriver.exe")
driver.get("https://destinationinsights.withgoogle.com/")
#listcountries = [{"name":"Worldwide","id":"WW"},{"name":"Afghanistan","id":"AF"},{"name":"Albania","id":"AL"},{"name":"Algeria","id":"DZ"},{"name":"\u00c5land Islands","id":"AX"},{"name":"American Samoa","id":"AS"},{"name":"Andorra","id":"AD"},{"name":"Angola","id":"AO"},{"name":"Anguilla","id":"AI"},{"name":"Antarctica","id":"AQ"},{"name":"Antigua and Barbuda","id":"AG"},{"name":"Argentina","id":"AR"},{"name":"Armenia","id":"AM"},{"name":"Aruba","id":"AW"},{"name":"Australia","id":"AU"}, {"name":"Austria","id":"AT"},{"name":"Azerbaijan","id":"AZ"},{"name":"Bahrain","id":"BH"},{"name":"Bangladesh","id":"BD"},{"name":"Barbados","id":"BB"},{"name":"Belarus","id":"BY"},{"name":"Belgium","id":"BE"},{"name":"Belize","id":"BZ"},{"name":"Benin","id":"BJ"},{"name":"Bermuda","id":"BM"},{"name":"Bhutan","id":"BT"},{"name":"Bolivia","id":"BO"},{"name":"Bosnia and Herzegovina","id":"BA"},{"name":"Botswana","id":"BW"},{"name":"Bouvet Island","id":"BV"},{"name":"Brazil","id":"BR"},{"name":"British Indian Ocean Territory","id":"IO"},{"name":"British Virgin Islands", "id":"VG"},{"name":"Brunei","id":"BN"},{"name":"Bulgaria","id":"BG"},{"name":"Burkina Faso","id":"BF"},{"name":"Burundi","id":"BI"},{"name":"Cambodia","id":"KH"},{"name":"Cameroon","id":"CM"},{"name":"Canada","id":"CA"},{"name":"Cape Verde","id":"CV"},{"name":"Caribbean Netherlands","id":"BQ"},{"name":"Cayman Islands","id":"KY"},{"name":"Central African Republic","id":"CF"},{"name":"Chad","id":"TD"},{"name":"Chile","id":"CL"},{"name":"China","id":"CN"},{"name":"Christmas Island","id":"CX"},{"name":"Cocos (Keeling) Islands","id":"CC"},{"name":"Colombia","id":"CO"},{"name":"Comoros", "id":"KM"},{"name":"Cook Islands","id":"CK"},{"name":"Costa Rica","id":"CR"},{"name":"C\u00f4te d'Ivoire","id":"CI"},{"name":"Croatia","id":"HR"},{"name":"Cuba","id":"CU"},{"name":"Cura\u00e7ao","id":"CW"},{"name":"Cyprus","id":"CY"},{"name":"Czechia","id":"CZ"},{"name":"Democratic Republic of the Congo","id":"CD"},{"name":"Denmark","id":"DK"},{"name":"Djibouti","id":"DJ"},{"name":"Dominica","id":"DM"},{"name":"Dominican Republic","id":"DO"},{"name":"Ecuador","id":"EC"},{"name":"Egypt","id":"EG"},{"name":"El Salvador","id":"SV"},{"name":"Equatorial Guinea","id":"GQ"}, {"name":"Eritrea","id":"ER"},{"name":"Estonia","id":"EE"},{"name":"Eswatini","id":"SZ"},{"name":"Ethiopia","id":"ET"},{"name":"Falkland Islands (Islas Malvinas)","id":"FK"},{"name":"Faroe Islands","id":"FO"},{"name":"Federated States of Micronesia","id":"FM"},{"name":"Fiji","id":"FJ"},{"name":"Finland","id":"FI"},{"name":"France","id":"FR"},{"name":"French Guiana","id":"GF"},{"name":"French Polynesia","id":"PF"},{"name":"French Southern Territories","id":"TF"},{"name":"Gabon","id":"GA"},{"name":"Georgia","id":"GE"},{"name":"Germany","id":"DE"}, {"name":"Ghana","id":"GH"},{"name":"Gibraltar","id":"GI"},{"name":"Greece","id":"GR"},{"name":"Greenland","id":"GL"},{"name":"Grenada","id":"GD"},{"name":"Guadeloupe","id":"GP"},{"name":"Guam","id":"GU"},{"name":"Guatemala","id":"GT"},{"name":"Guernsey","id":"GG"},{"name":"Guinea-Bissau","id":"GW"},{"name":"Guinea","id":"GN"},{"name":"Guyana","id":"GY"},{"name":"Haiti","id":"HT"},{"name":"Honduras","id":"HN"},{"name":"Hong Kong","id":"HK"},{"name":"Hungary","id":"HU"},{"name":"Iceland","id":"IS"},{"name":"India","id":"IN"},{"name":"Indonesia","id":"ID"},{"name":"Iran","id":"IR"}, {"name":"Iraq","id":"IQ"},{"name":"Ireland","id":"IE"},{"name":"Isle of Man","id":"IM"},{"name":"Israel","id":"IL"},{"name":"Italy","id":"IT"},{"name":"Jamaica","id":"JM"},{"name":"Japan","id":"JP"},{"name":"Jersey","id":"JE"},{"name":"Jordan","id":"JO"},{"name":"Kazakhstan","id":"KZ"},{"name":"Kenya","id":"KE"},{"name":"Kiribati","id":"KI"},{"name":"Kuwait","id":"KW"},{"name":"Kyrgyzstan","id":"KG"},{"name":"Laos","id":"LA"},{"name":"Latvia","id":"LV"},{"name":"Lebanon","id":"LB"},{"name":"Lesotho","id":"LS"},{"name":"Liberia","id":"LR"},{"name":"Libya","id":"LY"},{"name":"Liechtenstein", "id":"LI"},{"name":"Lithuania","id":"LT"},{"name":"Luxembourg","id":"LU"},{"name":"Macao","id":"MO"},{"name":"Madagascar","id":"MG"},{"name":"Malawi","id":"MW"},{"name":"Malaysia","id":"MY"},{"name":"Maldives","id":"MV"},{"name":"Mali","id":"ML"},{"name":"Malta","id":"MT"},{"name":"Marshall Islands","id":"MH"},{"name":"Martinique","id":"MQ"},{"name":"Mauritania","id":"MR"},{"name":"Mauritius","id":"MU"},{"name":"Mayotte","id":"YT"},{"name":"Mexico","id":"MX"},{"name":"Moldova","id":"MD"},{"name":"Monaco","id":"MC"},{"name":"Mongolia", "id":"MN"},{"name":"Montenegro","id":"ME"},{"name":"Montserrat","id":"MS"},{"name":"Morocco","id":"MA"},{"name":"Mozambique","id":"MZ"},{"name":"Myanmar (Burma)","id":"MM"},{"name":"Namibia","id":"NA"},{"name":"Nauru","id":"NR"},{"name":"Nepal","id":"NP"},{"name":"Netherlands","id":"NL"},{"name":"New Caledonia","id":"NC"},{"name":"New Zealand","id":"NZ"},{"name":"Nicaragua","id":"NI"},{"name":"Niger","id":"NE"},{"name":"Nigeria","id":"NG"},{"name":"Niue","id":"NU"},{"name":"Norfolk Island","id":"NF"},{"name":"North Korea","id":"KP"},{"name":"North Macedonia", "id":"MK"},{"name":"Northern Mariana Islands","id":"MP"},{"name":"Norway","id":"NO"},{"name":"Oman","id":"OM"},{"name":"Pakistan","id":"PK"},{"name":"Palau","id":"PW"},{"name":"Palestinian Territory","id":"PS"},{"name":"Panama","id":"PA"},{"name":"Papua New Guinea","id":"PG"},{"name":"Paraguay","id":"PY"},{"name":"Peru","id":"PE"},{"name":"Philippines","id":"PH"},{"name":"Pitcairn","id":"PN"},{"name":"Poland","id":"PL"},{"name":"Portugal","id":"PT"},{"name":"Puerto Rico","id":"PR"},{"name":"Qatar","id":"QA"},{"name":"Republic of the Congo","id":"CG"},{"name":"R\u00e9union", "id":"RE"},{"name":"Romania","id":"RO"},{"name":"Russia","id":"RU"},{"name":"Rwanda","id":"RW"},{"name":"Saint Barthelemy","id":"BL"},{"name":"Saint Helena, Ascension and Tristan da Cunha","id":"SH"},{"name":"Saint Kitts and Nevis","id":"KN"},{"name":"Saint Lucia","id":"LC"},{"name":"Saint Martin","id":"MF"},{"name":"Saint Pierre and Miquelon","id":"PM"},{"name":"Saint Vincent and the Grenadines","id":"VC"},{"name":"Samoa","id":"WS"},{"name":"San Marino","id":"SM"},{"name":"S\u00e3o Tom\u00e9 and Pr\u00edncipe", "id":"ST"},{"name":"Saudi Arabia","id":"SA"},{"name":"Senegal","id":"SN"},{"name":"Serbia","id":"RS"},{"name":"Seychelles","id":"SC"},{"name":"Sierra Leone","id":"SL"},{"name":"Singapore","id":"SG"},{"name":"Sint Maarten","id":"SX"},{"name":"Slovakia","id":"SK"},{"name":"Slovenia","id":"SI"},{"name":"Solomon Islands","id":"SB"},{"name":"Somalia","id":"SO"},{"name":"South Africa","id":"ZA"},{"name":"South Korea","id":"KR"},{"name":"South Sudan","id":"SS"},{"name":"Spain","id":"ES"},{"name":"Sri Lanka","id":"LK"},{"name":"Sudan","id":"SD"},{"name":"Suriname","id":"SR"},{"name":"Svalbard and Jan Mayen", "id":"SJ"},{"name":"Sweden","id":"SE"},{"name":"Switzerland","id":"CH"},{"name":"Syria","id":"SY"},{"name":"Taiwan","id":"TW"},{"name":"Tajikistan","id":"TJ"},{"name":"Tanzania","id":"TZ"},{"name":"Thailand","id":"TH"},{"name":"The Bahamas","id":"BS"},{"name":"The Gambia","id":"GM"},{"name":"Timor-Leste","id":"TL"},{"name":"Togo","id":"TG"},{"name":"Tokelau","id":"TK"},{"name":"Tonga","id":"TO"},{"name":"Trinidad and Tobago","id":"TT"},{"name":"Tunisia","id":"TN"},{"name":"Turkey","id":"TR"},{"name":"Turkmenistan","id":"TM"},{"name":"Turks and Caicos Islands","id":"TC"}, {"name":"Tuvalu","id":"TV"},{"name":"U.S. Virgin Islands","id":"VI"},{"name":"Uganda","id":"UG"},{"name":"Ukraine","id":"UA"},{"name":"United Arab Emirates","id":"AE"},{"name":"United Kingdom","id":"GB"},{"name":"United States","id":"US"},{"name":"Uruguay","id":"UY"},{"name":"Uzbekistan","id":"UZ"},{"name":"Vanuatu","id":"VU"},{"name":"Vatican","id":"VA"},{"name":"Venezuela","id":"VE"},{"name":"Vietnam","id":"VN"},{"name":"Wallis and Futuna","id":"WF"},{"name":"Yemen","id":"YE"},{"name":"Zambia","id":"ZM"},{"name":"Zimbabwe","id":"ZW"}]
destination_country = 'Spain'
origin_country = 'Morocco'
filepath = (r"google-insights.xlsx")

# ACCEPT COOKIES
driver.find_elements_by_xpath('/html/body/div[1]/div/span[2]/a[2]')[0].click()
sleep(2)

# SELECT  Origin country
driver.find_element_by_xpath('//*[@id="select_10"]').click() 
sleep(2) 

a = driver.find_element_by_xpath('//*[@id="select_listbox_12"]')
b= a.get_attribute("outerHTML")
soup = BeautifulSoup(b, 'lxml')
tag = soup.find_all('md-option')

for x in range(len(tag)):
    if origin_country in str(tag[x].text):
        driver.find_element_by_xpath('//*[@id="' + str(tag[x].get('id')) +'"]').click()  
sleep(3) 

# SELECT  Destination country
driver.find_element_by_xpath('/html/body/div[1]/div[6]/div/div/div[1]/div[2]/md-content[1]/md-select').click() 
sleep(2) 

a = driver.find_element_by_xpath('//*[@id="select_listbox_18"]')
b= a.get_attribute("outerHTML")
soup = BeautifulSoup(b, 'lxml')
tag = soup.find_all('md-option')

for x in range(len(tag)):
    if destination_country in str(tag[x].text):
        driver.find_element_by_xpath('//*[@id="' + str(tag[x].get('id')) +'"]').click()  
sleep(3) 

# Submit
driver.find_element_by_xpath('/html/body/div[1]/div[6]/div/div/div[2]/div[2]/button').click()
sleep(5)


listtotal = []

table1 = driver.find_element_by_xpath("/html/body/div[1]/div[21]/div[2]") 
table2 = driver.find_element_by_xpath("/html/body/div[1]/div[17]/div[2]/div[4]") 
table3 = driver.find_element_by_xpath("/html/body/div[1]/div[13]/div[2]/div[4]") 
table1clean = table1.get_attribute("outerHTML") 
table2clean = table2.get_attribute("outerHTML") 
table3clean = table3.get_attribute("outerHTML") 
table1soup = BeautifulSoup(table1clean, "lxml") 
table2soup = BeautifulSoup(table2clean, "lxml") 
table3soup = BeautifulSoup(table3clean, "lxml") 
mydivsregions = table1soup.find_all("span", {"class": "sources__name ng-binding"})
mydivsregionsnumber = table1soup.find_all("span", {"class": "sources__total glue-small-text ng-binding"}) 
growthregions = table2soup.find_all("span", {"class": "geographic-demand__name ng-binding"}) 
growthregionsnumber = table2soup.find_all("span", {"class": "growth-scale__label glue-small-text ng-binding"}) 
mydivsregionssources = table3soup.find_all("span", {"class": "geographic-demand__name ng-binding"}) 
mydivsregionssourcesnumbers = table3soup.find_all("span", {"class": "geographic-demand__total glue-small-text ng-binding"})


regionsources = [] 
for x in range(len(mydivsregionssources)):
    regionsources.append([mydivsregionssources[x].text,int(mydivsregionssourcesnumbers[x].text)]) 

growth = []
for x in range(len(growthregions)):
    growth.append([growthregions[x].text,growthregionsnumber[x].text]) 

regions = [] 
for x in range(len(mydivsregions)):
    regions.append([mydivsregions[x].text,int(mydivsregionsnumber[x].text)]) 
    
dictionarytotal = {origin_country+'-'+destination_country : [regionsources, growth, regions]} 
listtotal.append(dictionarytotal)


# save into XL
wb = openpyxl.Workbook() 
wb.save(filepath) 
for x in listtotal:
    for k,v in x.items():
        key_dict = k 
        value_dict = v 
        ws1 = wb.create_sheet("Sheet_A") 
        if len(str(k)) < 31:
            ws1.title = str(k) 
            ws = wb[str(k)] 
        else:
            ws1.title = str(k)[0:30] 
            ws = wb[str(k)[0:30]] 
counter = 1 
for y in value_dict: 
    counter2 = 8
    ws.cell(row=1, column=1).value = 'origin: ' + origin_country
    ws.cell(row=2, column=1).value = 'destination: ' + destination_country
    ws.cell(row=3, column=1).value = 'trip type'
    ws.cell(row=4, column=1).value = 'demand category'
    ws.cell(row=5, column=1).value = 'date range'
    ws.cell(row=7, column=1).value = "Top demand by destination city" 
    ws.cell(row=7, column=2).value = "Interest" 
    ws.cell(row=7, column=4).value = "TOP GROWTH BY DESTINATION CITY" 
    ws.cell(row=7, column=5).value = "Interest" 
    ws.cell(row=7, column=7).value = "TOP DEMAND BY ORIGIN COUNTRY" 
    ws.cell(row=7, column=8).value = "Interest" 
    for z in y: 
        ws.cell(row=counter2, column=counter).value = z[0] 
        ws.cell(row=counter2, column=counter+1).value = z[1] 
        counter2 = counter2 + 1 
    counter = counter + 3
del wb["Sheet"] 
wb.save(filepath) 
wb.close()

driver.quit()
